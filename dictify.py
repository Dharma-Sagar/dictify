from collections import defaultdict
from pathlib import Path

import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import pyewts

conv = pyewts.pyewts()


def load_dicts():
    dicts = defaultdict(dict)
    dict_path = Path('data/dictionaries/converted')
    for f in sorted(dict_path.glob('*.txt')):
        name = f.stem
        lines = f.read_text().split('\n')
        for line in lines:
            if '|' not in line:
                continue

            lemma, entry = line.split('|')
            dicts[lemma][name] = entry

    return dicts


def dictify_text(string):
    string = string.replace('\n', ' ')
    words = []
    for w in string.split(' '):
        words.append([w, {}])

    dicts = load_dicts()
    for num, word in enumerate(words):
        lemma = word[0].rstrip('་')
        defs = dicts[lemma]
        # filter
        words[num][1]['defs'] = defs

    for num, word in enumerate(words):
        if num <= 0:
            continue
        if num >= len(words):
            continue

        prev_lemma = words[num-1][0]
        cur_lemma = words[num][0]
        if cur_lemma.startswith('-'):
            words[num-1][0] = prev_lemma[:-1]
            words[num][0] = cur_lemma[1:]
        if cur_lemma.endswith('།'):
            if prev_lemma[-2] != 'ང' and prev_lemma.endswith('་'):
                words[num-1][0] = prev_lemma[:-1]
            words[num][0] += ' '

    return words


def generate_xlsx(title, content, out_file):
    font = 'Jomolhari'
    lemma_style = Font(font, size=18, bold=True)
    tib_style = Font(font, size=15)
    alignmnt = Alignment(horizontal="left", vertical="distributed", wrap_text=True)

    def fill_cell(sheet, row, col, string, font_style, height=None, width=None, link=None):
        sheet.cell(row=row, column=col).value = string
        sheet.cell(row=row, column=col).font = font_style
        sheet.cell(row=row, column=col).alignment = alignmnt
        if height:
            sheet.row_dimensions[row].height = height
        if width:
            sheet.column_dimensions[get_column_letter(col)].width = width
        if link:
            sheet.cell(row=row, column=col).hyperlink = link

    wb = Workbook()
    wb.remove(wb['Sheet'])
    ws = wb.create_sheet(title=title)
    row, col = 1, 1
    width, height = 50, 50
    for word, data in content:
        defs = select_defs(data['defs'])
        if defs:
            link = gen_link(word)
            fill_cell(ws, row, col, word, lemma_style, height=height, link=link)
            col += 1
            defs = defs.split('\n\n')
            for d in defs:
                fill_cell(ws, row, col, d, tib_style, width=width)
                col += 1
        else:
            fill_cell(ws, row, col, word, lemma_style, height=height)
            col += 1
        row += 1
        col = 1

    wb.save(out_file)


def select_defs(defs):
    cats = yaml.safe_load(Path('data/dictionaries/dict_cats.yaml').read_text())
    english, tibetan = cats['english']['dictionary'], cats['tibetan']['dictionary']

    selected = {}
    # selecting English definition
    for full, name in english:
        if full in defs:
            selected['en'] = (name, defs[full])
            break

    # selecting Tibetan definition
    for full, name in tibetan:
        if full in defs:
            selected['bo'] = (name, defs[full])
            break

    # format selected
    if 'en' in selected and 'bo' in selected:
        return f"{selected['en'][0]}\n{selected['en'][1]}\n\n{selected['bo'][0]}\n{selected['bo'][1]}"
    elif 'en' in selected:
        return f"{selected['en'][0]}\n{selected['en'][1]}"
    elif 'bo' in selected:
        return f"{selected['bo'][0]}\n{selected['bo'][1]}"
    else:
        return None


def gen_link(word):
    link_pattern = 'https://dictionary.christian-steinert.de/#%7B%22activeTerm%22%3A%22{word}%22%2C%22' \
                   'lang%22%3A%22tib%22%2C%22inputLang%22%3A%22tib%22%2C%22currentListTerm%22%3A%22{word}%22%2C%22' \
                   'forceLeftSideVisible%22%3Atrue%2C%22offset%22%3A0%7D'
    wylie = conv.toWylie(word)
    return link_pattern.format(word=wylie)


def generate_html(title, content, out_file):
    html_pattern = '<!doctype html><meta charset=utf-8><title>{title}</title>\n<html lang=en>\n' \
                '<head>\n<meta charset=utf-8>\n<title>{title}</title>\n</head><style>\n{style}\n</style>\n' \
                   '<body>\n<p>\n{body}\n</p>\n</body>\n</html>'
    entry_pattern = '<a href="{link}" target="_blank" title="{defs}">{word}</a>'
    style = """a:link {
  text-decoration: none;
  color: black;
}

a:visited {
  color: gray;
  text-decoration: none;
  display: inline-block;
}

a:hover {
  background-color: yellow;
}"""
    font_size = 'x-large'  # large, larger, medium, small, smaller, x-large, xx-large, xx-small, or "0cm" -> change value
    style += "p {\n    font-size: " + font_size + ";\n}"

    body = []
    for word, data in content:
        defnts = data['defs']
        if defnts:
            defs = select_defs(defnts)
            link = gen_link(word)
            def_out = entry_pattern.format(link=link, word=word, defs=defs)
            body.append(def_out)
        else:
            body.append(word)

    out = html_pattern.format(title=title, style=style, body=''.join(body))

    Path(out_file).write_text(out)


def batch_process(in_path):
    for f in Path(in_path).glob('*.txt'):
        dump = f.read_text()
        out = dictify_text(dump)
        generate_html(f.stem, out, Path('output') / (f.stem + '.html'))
        generate_xlsx(f.stem, out, Path('output') / (f.stem + '.xlsx'))


batch_process('input')
